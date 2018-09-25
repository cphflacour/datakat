import os
import pylons
import inspect
import losser.losser
import re
import json
import urllib2
import sys
import time
from contextlib import closing
import requests
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Style, Font
from openpyxl.styles import Border, Side
from openpyxl.writer.excel import save_virtual_workbook
from ckanapi.remoteckan import RemoteCKAN
import urllib2
import urlparse
import json
import shutil
import datetime
import urllib
from itertools import islice, izip
from ckan.logic import get_action
from ckanext.cphmetadata.plugin import get_quality_translation, get_frequency_translation
from ckan.controllers.admin import AdminController

def login(username, password):
    '''
    Login to CKAN.

    Returns a ``requests.Session`` instance with the CKAN
    session cookie.
    '''
    s = requests.Session()
    data = {'login': username, 'password': password}
    url = "http://metadata.bydata.dk" + '/login_generic'
    r = s.post(url, data=data)
    if 'field-login' in r.text:
        # Response still contains login form
        raise RuntimeError('Login failed.')
    return s

def download_resource_data(session,url):
    return session.get(url)

def fetchDistTables():
        ##########################################################
        #  Returns a list of objects with room for metadata
        ##########################################################
        # Initialize database connection to KGB DIST
        kgbdist_url = "http://wfs-kbhkort.spatialsuite.dk/k101/" \
                "ows?service=WFS&version=1.0.0" \
                "&request=GetFeature" \
                "&typeName=k101:ekst_table_upd_status" \
                "&outputFormat=json"

        response = urllib2.urlopen(kgbdist_url)
        rows = json.load(response)["features"]

        kgbobjects = []
        for row in rows:
	    object = {
        	        "scheme": row["properties"]["schema_name"],
	                "name": row["properties"]["table_name"],
        	        "metadata": None,
                	"bydata": None,
               		"ignore": False,
               		"qgis": False,
               		"color": None,
                	"note": "",
                	"fagkontakt": "",
            	}
            kgbobjects.append(object)

        return kgbobjects

def addMarks(distTables, ignoreExcel):
        # # # #
        # This code block updates the current information about contact persons and metadata for metadata in KGB
        # Based on a previously marked list
        # # # #
        
        wb = load_workbook(ignoreExcel)  # This is the content of the Database, but marked
        ws = wb.active

        wbDest = Workbook()
        wsDest = wbDest.active
        wsDest.title = "KGB Marked"
        # Add new occurences
        # We dont know which is larger, KGB or the list.
        maxRows = len(distTables) if len(distTables) >= ws.max_row else ws.max_row

        distNames = []
        for object in distTables:
	        distNames.append(object["name"])

        distNames = sorted(distNames)

        key = 1
        for row in range(1, maxRows):  # Iterate over all the tables in the marked list
	        key = key + 1
            	if ws.cell(row=key, column=1).value != '':  # Pre-cautionary empty-check
                    name = ws.cell(row=key, column=1).value

                    note(distTables, name, ws.cell(row=key, column=5).value, ws.cell(row=key, column=4).value)
                    color(distTables, name, ws.cell(row=key, column=1).fill.start_color)
                    if ws.cell(row=key, column=2).value == 'x':
                        # If we see the x or ? we know to ignore the data
                    	ignore(distTables, name)
                    	continue

                # Update the bydata name
                if ws.cell(row=key, column=3).value != '':  # Bydata navn
                    # bydata is a function that translates the names to emails etc.
                    bydata(distTables, name, ws.cell(row=key, column=3).value)

        return distTables

def addMetadata(distTables):

        APIKEYM = '0724a4f5-265f-4c63-9c8d-6c72f59424a1'
        siteM = 'http://metadata.bydata.dk'

        try:
            ckan = RemoteCKAN(siteM, apikey=APIKEYM)
        except:
            print("Ckan instance could not be initialized")

        datasetName = ''  # Find all dataset
        try:
            dataset = ckan.action.package_search(q=datasetName, rows=10000, include_private=True)
        except Exception as e:
            print(e)
            print("Could not find dataset, or something went wrong.. exiting")
            sys.exit()

        for i in dataset['results']:
            # metadata is a function that translates metadata to proper formatting and updates our objects
            metadata(distTables, i['name'], str("http://metadata.bydata.dk/dataset/" + i['name']))

        return distTables


def createReport(finalTables):
        wbDest = Workbook()
        wsDest = wbDest.active
        wsDest.title = "KGB Marked"

        key = 2
        c = wsDest["H2"]
        wsDest.freeze_panes = c
        fill = PatternFill("solid", fgColor="000080", fill_type="solid")
        for i in range(1, 8):
            wsDest.cell(row=1, column=i).style = Style(font=Font(bold=True, color='FFFFFF'), fill=fill)

        # Header
        wsDest.cell(row=1, column=1).value = "Datasaet navn"
        wsDest.cell(row=1, column=2).value = "medtages ikke"
        wsDest.cell(row=1, column=3).value = "bydata ansvarlig"
        wsDest.cell(row=1, column=4).value = "Fagkontakt"
        wsDest.cell(row=1, column=5).value = "Note"
        wsDest.cell(row=1, column=6).value = "Metadata"
        wsDest.cell(row=1, column=7).value = "Qgis"

        # Color explanation
        wsDest.cell(row=2, column=10).value = "Farveforklaring"
        wsDest.cell(row=3, column=10).value = "OK."
        wsDest.cell(row=3, column=10).fill = PatternFill(fgColor="72FF72", fill_type="solid")  # Green

        wsDest.cell(row=4, column=10).value = "Qgis el. Metadata mangler"
        wsDest.cell(row=4, column=10).fill = PatternFill(fgColor="FFEE72", fill_type="solid")

        wsDest.cell(row=5, column=10).value = "Qgis el. Metadata selvom \"x\""
        wsDest.cell(row=5, column=10).fill = PatternFill(fgColor="FFD072", fill_type="solid")

        wsDest.cell(row=6, column=10).value = "Nyt data"
        wsDest.cell(row=6, column=10).fill = PatternFill(fgColor="FF7272", fill_type="solid")

        wsDest.cell(row=7, column=10).value = "Ignoreret"
        wsDest.cell(row=7, column=10).fill = PatternFill(fgColor="D6D6D6", fill_type="solid")

        for object in sorted(finalTables, key=lambda x: (x['name'].lower(), x['name']),
                        reverse=False):  # Iterate over all the tables in KGB
            wsDest.cell(row=key, column=1).value = object["name"]
            for i in range(1, 9):  # Colors the entire row
                if (object["ignore"]):
                    if (object["metadata"] == None and object["qgis"] == False):  # Gray
                        fill = PatternFill(fgColor="D6D6D6", fill_type="solid")
                        wsDest.cell(row=key, column=i).fill = fill
                    else:
                        fill = PatternFill(fgColor="FFD072", fill_type="solid")  # Orange
                        wsDest.cell(row=key, column=i).fill = fill
                else:
                    if (object["qgis"] == False and object["metadata"] == None):  # Red
                        fill = PatternFill(fgColor="FF7272", fill_type="solid")
                        wsDest.cell(row=key, column=i).fill = fill
                    elif (object["qgis"] == False or object["metadata"] == None):  # Yellow
                        fill = PatternFill(fgColor="FFEE72", fill_type="solid")
                        wsDest.cell(row=key, column=i).fill = fill
                    else:
                        fill = PatternFill(fgColor="72FF72", fill_type="solid")  # Green
                        wsDest.cell(row=key, column=i).fill = fill
                border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
                wsDest.cell(row=key, column=i).border = border


            if (object["ignore"]):
                wsDest.cell(row=key, column=2).value = "x"
            else:
                wsDest.cell(row=key, column=2).value = ""

            if (object["bydata"]):
                wsDest.cell(row=key, column=3).value = object["bydata"]
            else:
                wsDest.cell(row=key, column=3).value = ""
            wsDest.cell(row=key, column=4).value = object["fagkontakt"]
            wsDest.cell(row=key, column=5).value = object["note"]
            wsDest.cell(row=key, column=6).value = object["metadata"]
            if (object["qgis"]):
                wsDest.cell(row=key, column=7).value = "x"
            else:
                wsDest.cell(row=key, column=7).value = ""

            key = key + 1

        for column_cells in wsDest.columns:
            try:
                length = min(40, max(len(str(cell.value)) for cell in column_cells) + 5)
                wsDest.column_dimensions[column_cells[0].column].width = length
	    except:
		wsDest.column_dimensions[column_cells[0].column].width = 40
        #resource_id = "29afd7bf-4ca4-4da7-bc89-f28e2b588119"


        #output = HttpResponse(mimetype='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #file_name = "Test.xlsx"
        #output['Content-Disposition'] = 'attachment; filename=' + file_name

        return save_virtual_workbook(wbDest)


def ignore(list, name):
        for x in list:
            if x["name"] == name:
                x["ignore"] = True

def bydata(list, name, bydata):
        for x in list:
            if x["name"] == name:
                if (bydata != None):
                    x["bydata"] = str(bydata).lower()
                else:
                    x["bydata"] = None

def metadata(list, name, metadata):
        for x in list:
            if x["name"] == name:
                x["metadata"] = metadata

def note(list, name, note, fagkontakt):
        for x in list:
            if x["name"] == name:
                x["note"] = note
                x["fagkontakt"] = fagkontakt

def color(list, name, color):
        for x in list:
            if x["name"] == name:
                x["color"] = color

def addQgis(distTables,qgisTables):

    for object in distTables:
        for table in qgisTables:
            if(object["scheme"]+"."+object["name"] == table[0]+"."+table[1]):
                object["qgis"] = True

    return distTables


def fetchQgisTables(qgisFile):

        file = qgisFile.getvalue()
        output = []

        p = re.compile(r'table=(\S+)', re.IGNORECASE)

        for line in file.splitlines():
            if "<layername>" in line or "<datasource>" in line:
                m = p.search(line)
                if m:
                    match = m.group(0).replace("table=","").replace("\"","")
                    output.append(match.split("."))

        return output


class ReportExportController(AdminController):
    """This controller exports datasets to a CSV file. Requires sysadmin role to download as private datasets are also included"""

    def _absolute_path(self, relative_path):
        """Return an absolute path given a path relative to this Python file."""
        return os.path.join(os.path.dirname(os.path.abspath(
            inspect.getfile(inspect.currentframe()))), relative_path)

    def download(self):
        """Uses package_search action to get all datasets in JSON format and transform to CSV"""

    	url_excel = "http://metadata.bydata.dk/dataset/4512eb02-0bb7-4524-8293-6fcffbd92762/resource/29afd7bf-4ca4-4da7-bc89-f28e2b588119/download/"
    	url_qgis = "http://metadata.bydata.dk/dataset/4512eb02-0bb7-4524-8293-6fcffbd92762/resource/ea61eb4f-702d-4af2-bac2-e7a11960482c/download/"

    	session = login('report_bot', 'Absalon2017')
    	data_mark = download_resource_data(session, url_excel)
    	data_qgis = download_resource_data(session, url_qgis)

    	with closing(data_mark), io.BytesIO(data_mark.content) as archive_excel:
            with closing(data_qgis), io.BytesIO(data_qgis.content) as archive_qgis:
                wb = load_workbook(archive_excel)  # This is the content of the Database, but marked
                ws = wb.active

                distTables = fetchDistTables()  # Fetches all the tables in DIST KGB
                qgisTables = fetchQgisTables(archive_qgis)  # Fetches all current tables and datasources in QGIS

                distTables = addMarks(distTables,archive_excel)  # This function reads the excel file, adds new occurences and marks "x"
                distTables = addMetadata(distTables)  # This function checks if the table has metadata

                finalTables = addQgis(distTables,qgisTables) #This function checks if the table is in QGIS

                file = createReport(finalTables)  # Will receive the name "kgb_report + "todays date" + .xlsx"
		

		filename = "kgb_report"+str(datetime.datetime.now().strftime("%d_%m_%Y"))+".xlsx"
                pylons.response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8' #content type for XLSX files (mime type)
                pylons.response.headers['Content-Disposition'] = 'attachment; filename='+filename

                return file
